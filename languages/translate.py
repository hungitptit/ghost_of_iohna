import openpyxl
import json
# Đường dẫn tới tệp Excel
excel_file_path = 'locale.xlsx'

# Tạo đối tượng Workbook từ tệp Excel
workbook = openpyxl.load_workbook(excel_file_path)

# Lấy sheet 
game_sheet_name = 'Game'
dialog_sheet_name = 'Dialog'
game_sheet = workbook[game_sheet_name]
dialog_sheet = workbook[dialog_sheet_name]

# Tạo các dictionary cho dữ liệu tiếng Anh và tiếng Việt
english_data = {}
vietnamese_data = {}

# Đọc dữ liệu từ các hàng trong sheet
for row in game_sheet.iter_rows(values_only=True):
    key, en, vi = row
    english_data[game_sheet_name + '.' + key] = en
    vietnamese_data[game_sheet_name + '.' + key] = vi

for row in dialog_sheet.iter_rows(values_only=True):
    key, en, vi = row
    english_data[dialog_sheet_name + '.' + str(key)] = en
    vietnamese_data[dialog_sheet_name + '.' + str(key)] = vi

# Xuất ra file JSON tiếng Anh
with open('English/main.json', 'w') as english_file:
    json.dump(english_data, english_file, indent=4)

# Xuất ra file JSON tiếng Việt
with open('Vietnamese/main.json', 'w', encoding='utf-8') as vietnamese_file:
    json.dump(vietnamese_data, vietnamese_file, ensure_ascii=False, indent=4)

print('Exported JSON files successfully!')